"""CSV + XLSX writers. Column order is locked by §12.1."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

# Exact column order per PLAN.md §12.1. Do NOT reorder without updating the
# spec and the Google Sheets writer (which consumes the same list).
COLUMNS: list[str] = [
    "rank",
    "overall_score",
    "thesis_fit",
    "team_quality",
    "earliness",
    "traction",
    "contactability",
    "name",
    "website",
    "one_liner",
    "sector",
    "sub_sector",
    "hq_city",
    "hq_country",
    "founded_year",
    "stage_signal",
    "team_size_estimate",
    "founder_names",
    "founder_linkedins",
    "founder_emails",
    "traction_signals",
    "funding_signals",
    "sources",
    "source_urls",
    "rationale_de",
    "red_flags",
    "discovered_at",
    "last_enriched",
    "scored_at",
    "canonical_id",
]

LIST_FIELDS = {
    "founder_names",
    "founder_linkedins",
    "founder_emails",
    "traction_signals",
    "funding_signals",
    "sources",
    "source_urls",
    "red_flags",
}

LIST_SEPARATOR = "|"


def write_csv(rows: list[dict], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=COLUMNS)
    df.to_csv(path, index=False)
    return path


def write_xlsx(rows: list[dict], path: Path) -> Path:
    """Write XLSX with bold header row + frozen header + rank column."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=COLUMNS)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="leads")
        ws = writer.sheets["leads"]
        # Freeze row 1 + column A so the rank stays visible while scrolling.
        ws.freeze_panes = "B2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        # Sensible default column widths — wider for text-heavy columns.
        text_heavy = {
            "name", "one_liner", "long_description", "rationale_de",
            "traction_signals", "funding_signals", "source_urls",
            "founder_emails", "founder_linkedins",
        }
        for idx, col in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = (
                40 if col in text_heavy else 16
            )
    return path
