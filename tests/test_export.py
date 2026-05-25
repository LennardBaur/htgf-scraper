"""End-to-end exporter tests — no network, no Google Sheets."""

from __future__ import annotations

import csv as csvlib
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from htgf_sourcer import db
from htgf_sourcer.enrich import canonical_id_for
from htgf_sourcer.export import (
    assemble_rows,
    export_all,
    write_run_summary,
)
from htgf_sourcer.exporters.csv_writer import COLUMNS
from htgf_sourcer.exporters.markdown_onepager import slugify, write_onepager
from htgf_sourcer.fetch import normalize_url
from htgf_sourcer.models import (
    EnrichedStartup,
    Founder,
    RawLead,
    Score,
    Source,
    Stage,
)

# ---- fixtures ------------------------------------------------------------


@pytest.fixture
def seeded_db(tmp_path: Path) -> Path:
    """A DB with two enriched startups, one scored. Returns the path."""
    p = tmp_path / "state.db"
    db.init_db(p)

    enriched_a = EnrichedStartup(
        canonical_id=canonical_id_for(normalize_url("https://acme.de/")),
        name="Acme GmbH",
        website="https://acme.de/",
        one_liner="AI for the Mittelstand",
        long_description="Acme builds AI tools for German Mittelstand operations.",
        sector="B2B SaaS",
        sub_sector="OperationsAI",
        hq_city="Berlin",
        hq_country="DE",
        founded_year=2024,
        stage_signal=Stage.PRE_SEED,
        team_size_estimate=8,
        founders=[
            Founder(
                name="Ada Lovelace",
                role="CEO",
                linkedin_url="https://linkedin.com/in/ada",
                email="ada@acme.de",
                background="PhD CS, prior exit",
            )
        ],
        traction_signals=["Pilot with Siemens", "10k GitHub stars"],
        funding_signals=["EXIST 2025"],
        sources=[Source.HACKERNEWS, Source.GITHUB],
        source_urls=["https://acme.de/", "https://github.com/acme"],
        last_enriched=datetime(2026, 5, 1, 10, 0, 0),
    )

    enriched_b_unscored = EnrichedStartup(
        canonical_id=canonical_id_for(normalize_url("https://beta.de/")),
        name="Beta Werkstatt",
        website="https://beta.de/",
        one_liner="A second startup, no score yet",
        long_description="Beta does cool things but hasn't been scored.",
        sector="DevTools",
        sources=[Source.BETALIST],
        source_urls=["https://beta.de/"],
        last_enriched=datetime(2026, 5, 2, 10, 0, 0),
    )

    score_a = Score(
        canonical_id=enriched_a.canonical_id,
        thesis_fit=5,
        team_quality=4,
        earliness=4,
        traction=3,
        contactability=4,
        overall=4.30,
        rationale="Starkes Team mit klarem B2B-SaaS-Fokus in DACH.",
        red_flags=["Frühe Phase, Umsatz noch unbestätigt"],
        scored_at=datetime(2026, 5, 5, 12, 0, 0),
    )

    lead = RawLead(
        source=Source.HACKERNEWS,
        source_id="hn-1",
        name="Acme GmbH",
        website="https://acme.de/",
        discovered_at=datetime(2026, 4, 1, 8, 0, 0),
    )

    with db.connect(p) as conn:
        db.upsert_enriched(conn, enriched_a)
        db.upsert_enriched(conn, enriched_b_unscored)
        db.upsert_score(conn, score_a)
        db.upsert_lead(conn, "lead-1", lead)

    return p


# ---- assemble_rows -------------------------------------------------------


def test_assemble_rows_ranks_scored_first_unscored_last(seeded_db: Path):
    rows, scored_items = assemble_rows(db_path=seeded_db)
    assert len(rows) == 2
    # Acme is scored → rank 1 first; Beta unscored → blank rank, last.
    assert rows[0]["name"] == "Acme GmbH"
    assert rows[0]["rank"] == 1
    assert rows[0]["overall_score"] == pytest.approx(4.30)
    assert rows[1]["name"] == "Beta Werkstatt"
    assert rows[1]["rank"] == ""

    assert len(scored_items) == 1
    assert scored_items[0][0].name == "Acme GmbH"


def test_assemble_rows_backfills_discovered_at(seeded_db: Path):
    rows, _ = assemble_rows(db_path=seeded_db)
    acme = next(r for r in rows if r["name"] == "Acme GmbH")
    # Came from the lead seeded with 2026-04-01.
    assert acme["discovered_at"].startswith("2026-04-01")


def test_row_dict_uses_pipe_separator_for_list_fields(seeded_db: Path):
    rows, _ = assemble_rows(db_path=seeded_db)
    acme = next(r for r in rows if r["name"] == "Acme GmbH")
    assert acme["sources"] == "hackernews|github"
    assert "Pilot with Siemens" in acme["traction_signals"]
    assert "|" in acme["traction_signals"]
    assert acme["founder_emails"] == "ada@acme.de"


# ---- CSV + XLSX ----------------------------------------------------------


def test_export_all_writes_csv_with_expected_columns(seeded_db: Path, tmp_path: Path):
    out = export_all(db_path=seeded_db, output_dir=tmp_path / "outputs", no_sheets=True)
    csv_path = out["csv"]
    with csv_path.open() as f:
        reader = csvlib.reader(f)
        header = next(reader)
        body = list(reader)
    assert header == COLUMNS
    assert len(body) == 2


def test_export_all_writes_xlsx_with_frozen_panes(seeded_db: Path, tmp_path: Path):
    out = export_all(db_path=seeded_db, output_dir=tmp_path / "outputs", no_sheets=True)
    wb = load_workbook(out["xlsx"])
    ws = wb.active
    assert ws.title == "leads"
    assert ws.freeze_panes == "B2"
    header = [cell.value for cell in ws[1]]
    assert header == COLUMNS
    # Bold header.
    assert ws["A1"].font.bold is True


# ---- one-pagers ----------------------------------------------------------


def test_export_writes_one_pager_only_for_scored(seeded_db: Path, tmp_path: Path):
    export_all(db_path=seeded_db, output_dir=tmp_path / "outputs", no_sheets=True)
    onepager_dir = tmp_path / "outputs" / "onepagers"
    files = list(onepager_dir.glob("*.md"))
    assert len(files) == 1
    content = files[0].read_text()
    # German labels present:
    for needle in (
        "**Rang:**",
        "## In einem Satz",
        "## Was sie machen",
        "## Gründer",
        "## Bewertung (LLM)",
        "**Begründung:**",
        "## Quellen",
        "Pre-Seed",  # stage label localized
    ):
        assert needle in content, f"missing in one-pager: {needle}"


def test_one_pager_filename_uses_rank_and_slug(tmp_path: Path):
    startup = EnrichedStartup(
        canonical_id="x" * 16,
        name="Müller & Söhne",
        one_liner="x",
        long_description="y",
        sector="B2B SaaS",
        last_enriched=datetime.utcnow(),
    )
    score = Score(
        canonical_id=startup.canonical_id,
        thesis_fit=3, team_quality=3, earliness=3, traction=3, contactability=3,
        overall=3.0,
        rationale="x",
        scored_at=datetime.utcnow(),
    )
    path = write_onepager(startup, score, rank=3, total=10, out_dir=tmp_path)
    assert path.name.startswith("03_")
    # umlauts folded to ascii
    assert "muller" in path.name.lower()


def test_slugify_handles_empty_and_unicode():
    assert slugify("") == "startup"
    assert slugify("Ümlaut Co!") == "umlaut-co"
    assert slugify("AAAAA" * 20).count("a") <= 50


# ---- run_summary.md ------------------------------------------------------


def test_run_summary_lists_top_and_sources(seeded_db: Path, tmp_path: Path):
    rows, _ = assemble_rows(db_path=seeded_db)
    path = write_run_summary(rows, db_path=seeded_db, path=tmp_path / "run_summary.md")
    text = path.read_text()
    assert "Run Summary" in text
    assert "Acme GmbH" in text
    assert "hackernews" in text
    # cost line present (zero in this fixture):
    assert "LLM cost" in text


# ---- google sheets graceful skip ----------------------------------------


def test_export_skips_sheets_when_creds_missing(seeded_db: Path, tmp_path: Path, monkeypatch):
    monkeypatch.delenv("GOOGLE_SA_PATH", raising=False)
    monkeypatch.delenv("GOOGLE_SHEET_ID", raising=False)
    out = export_all(db_path=seeded_db, output_dir=tmp_path / "outputs", no_sheets=False)
    assert "skipped" in out["sheets"].lower()
